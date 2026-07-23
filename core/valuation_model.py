"""valuation_model.py — Build downloadable Excel DCF/comps models from analysis results."""

from __future__ import annotations

import io
import math
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
_SUBHEADER_FONT = Font(name="Calibri", bold=True, color="E5E7EB", size=10)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F2937")
_LABEL_FONT = Font(name="Calibri", bold=True, size=10, color="1F2937")
_VALUE_FONT = Font(name="Calibri", size=10)
_HIGHLIGHT_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
_AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# ── Helpers ─────────────────────────────────────────────────────────────────


def _style_header_row(ws, row: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _write_label(ws, row: int, col: int, text: str, indent: int = 0):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _LABEL_FONT if indent == 0 else Font(name="Calibri", size=10, color="4B5563")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    cell.border = _THIN_BORDER
    return cell


def _write_value(ws, row: int, col: int, value: float | None, fmt: str = "#,##0"):
    cell = ws.cell(row=row, column=col)
    if value is not None and math.isfinite(value):
        cell.value = round(value, 2)
        cell.number_format = fmt
    else:
        cell.value = "—"
    cell.font = _VALUE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _THIN_BORDER
    return cell


def _write_pct(ws, row: int, col: int, value: float | None):
    if value is not None and math.isfinite(value):
        return _write_value(ws, row, col, value / 100, "0.0%")
    return _write_value(ws, row, col, None, "0.0%")


def _auto_width(ws, max_col: int, max_row: int, min_width: int = 12, max_width: int = 40):
    for col in range(1, max_col + 1):
        best = min_width
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                best = max(best, min(len(str(val)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = best


# ── Builder ─────────────────────────────────────────────────────────────────


class ValuationModelBuilder:
    """Build a multi-sheet Excel valuation workbook from analysis results."""

    def __init__(self, analysis: dict[str, Any], financials: dict[str, Any] | None = None):
        self.analysis = analysis
        self.financials = financials or {}
        self.wb = openpyxl.Workbook()
        self._build_summary()
        self._build_dcf()
        self._build_comps()
        self._build_assumptions()

    def build(self) -> io.BytesIO:
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    # ── Sheet 1: Summary ──────────────────────────────────────────────────

    def _build_summary(self):
        ws = self.wb.active
        assert ws is not None
        ws.title = "Summary"
        ws.sheet_properties.tabColor = "6366F1"

        scores = self.analysis.get("scores", {})
        project = self.analysis.get("project", {})
        adv = self.analysis.get("competitive_advantage", {})
        company = project.get("company", "Company")
        ticker = project.get("ticker", "N/A")

        ws.cell(
            row=1, column=1, value=f"Valuation Summary — {company} ({ticker})"
        ).font = _TITLE_FONT
        ws.cell(
            row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ).font = Font(name="Calibri", size=10, color="6B7280")

        # Score summary
        row = 4
        ws.cell(row=row, column=1, value="Dimension Scores (0–100)").font = _LABEL_FONT
        row += 1
        score_map = [
            ("Overall Growth", scores.get("overall_growth_score")),
            ("Moat Sustainability", scores.get("moat_sustainability")),
            ("Growth Capacity", scores.get("growth_capacity")),
            ("Execution Quality", scores.get("execution_quality")),
            ("Financial Resilience", scores.get("financial_resilience")),
            ("Risk Pressure", scores.get("risk_pressure")),
        ]
        _style_header_row(ws, row, 3)
        ws.cell(row=row, column=1, value="Dimension")
        ws.cell(row=row, column=2, value="Score")
        ws.cell(row=row, column=3, value="Rating")
        row += 1
        for dim, val in score_map:
            _write_label(ws, row, 1, dim)
            _write_value(ws, row, 2, val, "0")
            rating = self._rating_for_score(val)
            cell = ws.cell(row=row, column=3, value=rating)
            cell.font = _VALUE_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
            if rating == "Strong":
                cell.fill = _GREEN_FILL
            elif rating == "Watch":
                cell.fill = _AMBER_FILL
            row += 1

        # Competitive advantage
        row += 1
        ws.cell(row=row, column=1, value="Competitive Advantage").font = _LABEL_FONT
        row += 1
        _write_label(ws, row, 1, "Rating")
        ws.cell(row=row, column=2, value=adv.get("rating", "—")).font = _VALUE_FONT
        row += 1
        _write_label(ws, row, 1, "Assessment")
        ws.cell(row=row, column=2, value=adv.get("assessment", "—")).font = _VALUE_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        if scores:
            row += 2
            ws.cell(row=row, column=1, value="Investment Rating").font = _LABEL_FONT
            overall = scores.get("overall_growth_score", 50)
            rating, color = self._investment_rating(overall)
            cell = ws.cell(row=row, column=2, value=f"{rating} ({overall:.0f}/100)")
            cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        _auto_width(ws, 5, row)

    # ── Sheet 2: DCF Model ────────────────────────────────────────────────

    def _build_dcf(self):
        ws = self.wb.create_sheet("DCF Model")
        ws.sheet_properties.tabColor = "10B981"

        project = self.analysis.get("project", {})
        scores = self.analysis.get("scores", {})
        company = project.get("company", "Company")
        ticker = project.get("ticker", "N/A")

        ws.cell(row=1, column=1, value=f"DCF Valuation — {company} ({ticker})").font = _TITLE_FONT

        # WACC assumptions
        row = 3
        ws.cell(row=row, column=1, value="WACC Calculation").font = _LABEL_FONT
        row += 1
        wacc_items = [
            ("Risk-Free Rate", 0.035, "0.0%"),
            ("Equity Risk Premium", 0.08, "0.0%"),
            ("Beta", 1.0, "0.00"),
            ("Cost of Equity", None, "0.0%"),
            ("Pre-Tax Cost of Debt", 0.055, "0.0%"),
            ("Tax Rate", 0.20, "0.0%"),
            ("After-Tax Cost of Debt", None, "0.0%"),
            ("Debt / Total Capital", 0.30, "0.0%"),
            ("Equity / Total Capital", 0.70, "0.0%"),
            ("WACC", None, "0.0%"),
        ]
        _style_header_row(ws, row, 3)
        ws.cell(row=row, column=1, value="Component")
        ws.cell(row=row, column=2, value="Value")
        ws.cell(row=row, column=3, value="Notes")
        row += 1
        for label, default, fmt in wacc_items:
            _write_label(ws, row, 1, label)
            if default is not None:
                _write_pct(ws, row, 2, default) if "%" in fmt else _write_value(
                    ws, row, 2, default, fmt
                )
            else:
                _write_value(ws, row, 2, None)
            row += 1

        # Projection assumptions
        row += 1
        ws.cell(row=row, column=1, value="Projection Assumptions").font = _LABEL_FONT
        row += 1
        proj_items = [
            ("Revenue CAGR (Yr 1–5)", 0.08, "0.0%"),
            ("Terminal Growth Rate", 0.025, "0.0%"),
            ("Target Operating Margin", 0.20, "0.0%"),
            ("Reinvestment Rate", 0.15, "0.0%"),
            ("Projection Years", 5, "0"),
        ]
        _style_header_row(ws, row, 3)
        ws.cell(row=row, column=1, value="Assumption")
        ws.cell(row=row, column=2, value="Value")
        ws.cell(row=row, column=3, value="Source")
        row += 1
        for label, default, fmt in proj_items:
            _write_label(ws, row, 1, label)
            if isinstance(default, float):
                _write_pct(ws, row, 2, default)
            else:
                _write_value(ws, row, 2, float(default), fmt)
            row += 1

        # Projected cash flows (simplified illustration)
        row += 1
        ws.cell(row=row, column=1, value="Projected Free Cash Flows").font = _LABEL_FONT
        row += 1
        _style_header_row(ws, row, 8)
        headers = ["", "Yr 1", "Yr 2", "Yr 3", "Yr 4", "Yr 5", "Terminal", "Total"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        row += 1

        # Use the growth rating score to drive illustrative projections
        base_fcf = 1000  # placeholder — would come from financial surface
        if self.financials:
            fcf_vals = self.financials.get("derived_metrics", {}).get("fcf", [base_fcf])
            base_fcf = fcf_vals[-1] if fcf_vals else base_fcf

        growth_rate = 0.08
        terminal_growth = 0.025
        wacc = 0.095  # illustrative

        cf_rows = [
            ("Free Cash Flow", [base_fcf * (1 + growth_rate) ** y for y in range(1, 6)]),
            ("Discount Factor", [1 / (1 + wacc) ** y for y in range(1, 6)]),
            ("PV of FCF", None),
            ("Sum of PV (Yr 1–5)", None),
            ("Terminal Value", None),
            ("PV of Terminal Value", None),
            ("Enterprise Value", None),
            ("(+) Cash", None),
            ("(−) Debt", None),
            ("Equity Value", None),
            ("Shares Outstanding", None),
            ("Fair Value Per Share", None),
        ]

        for label, cf_vals in cf_rows:
            _write_label(ws, row, 1, label)
            if label == "Free Cash Flow":
                wacc_items
                for y in range(5):
                    _write_value(ws, row, 2 + y, cf_vals[y], "#,##0")
            elif label == "Discount Factor":
                for y in range(5):
                    _write_value(ws, row, 2 + y, cf_vals[y], "0.0000")
            elif label == "PV of FCF":
                pvs = [base_fcf * (1 + growth_rate) ** y / (1 + wacc) ** y for y in range(1, 6)]
                for y in range(5):
                    _write_value(ws, row, 2 + y, pvs[y], "#,##0")
                _write_value(ws, row, 7, sum(pvs), "#,##0")
            elif label == "Sum of PV (Yr 1–5)":
                pvs = [base_fcf * (1 + growth_rate) ** y / (1 + wacc) ** y for y in range(1, 6)]
                _write_value(ws, row, 2, sum(pvs), "#,##0")
            elif label == "Terminal Value":
                pv_yr5 = base_fcf * (1 + growth_rate) ** 5
                tv = pv_yr5 * (1 + terminal_growth) / (wacc - terminal_growth)
                _write_value(ws, row, 2, tv, "#,##0")
            elif label == "PV of Terminal Value":
                pv_yr5 = base_fcf * (1 + growth_rate) ** 5
                tv = pv_yr5 * (1 + terminal_growth) / (wacc - terminal_growth)
                _write_value(ws, row, 2, tv / (1 + wacc) ** 5, "#,##0")
            elif label == "Enterprise Value":
                pvs = [base_fcf * (1 + growth_rate) ** y / (1 + wacc) ** y for y in range(1, 6)]
                pv_yr5 = base_fcf * (1 + growth_rate) ** 5
                tv = pv_yr5 * (1 + terminal_growth) / (wacc - terminal_growth)
                ev = sum(pvs) + tv / (1 + wacc) ** 5
                _write_value(ws, row, 2, ev, "#,##0").fill = _HIGHLIGHT_FILL
            elif label == "(+) Cash":
                _write_value(ws, row, 2, 500, "#,##0")
            elif label == "(−) Debt":
                _write_value(ws, row, 2, 2000, "#,##0")
            elif label == "Equity Value":
                pvs = [base_fcf * (1 + growth_rate) ** y / (1 + wacc) ** y for y in range(1, 6)]
                tv = (
                    base_fcf
                    * (1 + growth_rate) ** 5
                    * (1 + terminal_growth)
                    / (wacc - terminal_growth)
                )
                ev = sum(pvs) + tv / (1 + wacc) ** 5
                _write_value(ws, row, 2, ev + 500 - 2000, "#,##0").fill = _GREEN_FILL
            elif label == "Shares Outstanding":
                _write_value(ws, row, 2, 100, "#,##0")
            elif label == "Fair Value Per Share":
                pvs = [base_fcf * (1 + growth_rate) ** y / (1 + wacc) ** y for y in range(1, 6)]
                tv = (
                    base_fcf
                    * (1 + growth_rate) ** 5
                    * (1 + terminal_growth)
                    / (wacc - terminal_growth)
                )
                ev = sum(pvs) + tv / (1 + wacc) ** 5
                fv = (ev + 500 - 2000) / 100
                _write_value(ws, row, 2, fv, "#,##0.00").fill = _GREEN_FILL

            row += 1

        # Sensitivity table (simple 2-way)
        row += 2
        ws.cell(
            row=row, column=1, value="Sensitivity: Fair Value vs WACC × Growth Rate"
        ).font = _LABEL_FONT
        row += 1
        wacc_range = [0.085, 0.090, 0.095, 0.100, 0.105]
        growth_range = [0.015, 0.020, 0.025, 0.030, 0.035]
        ws.cell(row=row, column=1, value="WACC ↓ / Growth →")
        for j, g in enumerate(growth_range):
            ws.cell(row=row, column=2 + j, value=f"{g:.1%}")
        _style_header_row(ws, row, 1 + len(growth_range))
        row += 1
        for wi, w in enumerate(wacc_range):
            _write_label(ws, row, 1, f"{w:.1%}")
            for gi, g in enumerate(growth_range):
                pv_yr5 = base_fcf * (1 + growth_rate) ** 5
                tv = pv_yr5 * (1 + g) / (w - g)
                ev = (
                    sum([base_fcf * (1 + growth_rate) ** y / (1 + w) ** y for y in range(1, 6)])
                    + tv / (1 + w) ** 5
                )
                fv = (ev + 500 - 2000) / 100
                _write_value(ws, row, 2 + gi, fv, "#,##0.00")
            row += 1

        _auto_width(ws, 8, row)

    # ── Sheet 3: Comps Analysis ───────────────────────────────────────────

    def _build_comps(self):
        ws = self.wb.create_sheet("Comps Analysis")
        ws.sheet_properties.tabColor = "3B82F6"

        project = self.analysis.get("project", {})
        company = project.get("company", "Company")
        ticker = project.get("ticker", "N/A")

        ws.cell(
            row=1, column=1, value=f"Comparable Companies Analysis — {company} ({ticker})"
        ).font = _TITLE_FONT

        # Peer placeholders (would be populated from peer_analysis in Phase 4)
        row = 3
        ws.cell(row=row, column=1, value="Trading Multiples").font = _LABEL_FONT
        row += 1
        _style_header_row(ws, row, 8)
        headers = ["Company", "Ticker", "Market Cap", "EV", "Revenue", "EBITDA", "P/E", "EV/EBITDA"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        row += 1

        # Peer data with focus company first
        peer_data = [
            (company, ticker, None, None, None, None, None, None),
            ("Peer 1", "—", None, None, None, None, None, None),
            ("Peer 2", "—", None, None, None, None, None, None),
            ("Peer 3", "—", None, None, None, None, None, None),
        ]

        for p_name, p_ticker, *vals in peer_data:
            _write_label(ws, row, 1, p_name)
            _write_value(ws, row, 2, 0)
            for vi, v in enumerate(vals, 3):
                _write_value(ws, row, vi, v)
            if p_name == company:
                for c in range(1, 9):
                    ws.cell(row=row, column=c).fill = _HIGHLIGHT_FILL
            row += 1

        # Medians
        row += 1
        _write_label(ws, row, 1, "Median (excl. target)")
        for c in range(3, 9):
            _write_value(ws, row, c, None)
        for c in range(1, 9):
            ws.cell(row=row, column=c).fill = _SUBHEADER_FILL
            ws.cell(row=row, column=c).font = _SUBHEADER_FONT

        _auto_width(ws, 8, row)

    # ── Sheet 4: Assumptions ──────────────────────────────────────────────

    def _build_assumptions(self):
        ws = self.wb.create_sheet("Assumptions")
        ws.sheet_properties.tabColor = "8B5CF6"

        ws.cell(
            row=1, column=1, value="Valuation Assumptions — Editable Parameters"
        ).font = _TITLE_FONT
        ws.cell(
            row=2, column=1, value="Change values here and refresh DCF/Comps sheets."
        ).font = Font(name="Calibri", italic=True, size=10, color="6B7280")

        row = 4
        ws.cell(row=row, column=1, value="WACC Components").font = _LABEL_FONT
        row += 1
        items = [
            ("rf", "Risk-Free Rate", 3.5, "%"),
            ("erp", "Equity Risk Premium", 8.0, "%"),
            ("beta", "Beta", 1.0, ""),
            ("kd", "Pre-Tax Cost of Debt", 5.5, "%"),
            ("tax", "Tax Rate", 20.0, "%"),
            ("debt_pct", "Debt / Total Capital", 30.0, "%"),
        ]
        _style_header_row(ws, row, 4)
        for i, h in enumerate(["Key", "Component", "Value", "Unit"], 1):
            ws.cell(row=row, column=i, value=h)
        row += 1
        for key, label, val, unit in items:
            _write_label(ws, row, 1, key)
            _write_label(ws, row, 2, label)
            _write_value(ws, row, 3, val, "0.00")
            _write_label(ws, row, 4, unit)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Growth Drivers").font = _LABEL_FONT
        row += 1
        g_items = [
            ("rev_cagr", "Revenue CAGR (Yr 1–5)", 8.0, "%"),
            ("term_g", "Terminal Growth Rate", 2.5, "%"),
            ("op_margin", "Target Operating Margin", 20.0, "%"),
            ("reinvest", "Reinvestment Rate", 15.0, "%"),
        ]
        _style_header_row(ws, row, 4)
        for i, h in enumerate(["Key", "Assumption", "Value", "Unit"], 1):
            ws.cell(row=row, column=i, value=h)
        row += 1
        for key, label, val, unit in g_items:
            _write_label(ws, row, 1, key)
            _write_label(ws, row, 2, label)
            _write_value(ws, row, 3, val, "0.00")
            _write_label(ws, row, 4, unit)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Financial Summary").font = _LABEL_FONT
        row += 1
        if self.financials:
            is_items = {
                r["label"]: r["values"] for r in self.financials.get("income_statement", [])
            }
            bs_items = {r["label"]: r["values"] for r in self.financials.get("balance_sheet", [])}
            fm_items = self.financials.get("derived_metrics", {})

            label_map = {
                "doanh_thu_thuan": "Revenue (Last)",
                "loi_nhuan_sau_thue": "Net Income (Last)",
                "ebitda": "EBITDA (Last)",
            }
            for key, display in label_map.items():
                vals = is_items.get(key, [])
                if vals:
                    _write_label(ws, row, 1, display)
                    _write_value(ws, row, 2, vals[-1], "#,##0")
                    row += 1

            for metric, display in [
                ("gross_margin", "Gross Margin (%)"),
                ("net_margin", "Net Margin (%)"),
            ]:
                vals = fm_items.get(metric, [])
                if vals:
                    _write_label(ws, row, 1, display)
                    _write_value(ws, row, 2, vals[-1], "0.0")
                    row += 1
        else:
            ws.cell(
                row=row,
                column=1,
                value="Upload a report and run analysis to populate financial data.",
            ).font = Font(name="Calibri", italic=True, size=10, color="9CA3AF")

        _auto_width(ws, 4, row)

    # ── Rating helpers ─────────────────────────────────────────────────────

    @staticmethod
    def _rating_for_score(score: float | None) -> str:
        if score is None:
            return "—"
        if score >= 75:
            return "Strong"
        if score >= 60:
            return "Average"
        return "Watch"

    @staticmethod
    def _investment_rating(overall: float) -> tuple[str, str]:
        if overall >= 80:
            return ("BUY", "10B981")
        if overall >= 65:
            return ("ACCUMULATE", "34D399")
        if overall >= 50:
            return ("HOLD", "FBBF24")
        if overall >= 35:
            return ("REDUCE", "F97316")
        return ("SELL", "EF4444")
